from django.shortcuts import render
from django.db.models import Q
import django.contrib.auth.password_validation as validators
from django.core import exceptions
from django.db import transaction
from rest_framework.authtoken.models import Token
import json
import smtplib
from allauth.socialaccount.providers.google.views import GoogleOAuth2Adapter
from allauth.socialaccount.providers.facebook.views import FacebookOAuth2Adapter
from dj_rest_auth.registration.views import SocialLoginView
from allauth.socialaccount.providers.oauth2.client import OAuth2Client
from django.core.files import File
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from django.core.serializers import serialize
import os, shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import mimetypes
from rest_framework import generics
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.authentication import TokenAuthentication
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from . models import *
from django.contrib.sites.shortcuts import get_current_site
from io import BytesIO
from django.template.loader import get_template, render_to_string
from xhtml2pdf import pisa
from rest_framework import filters
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response 
from django_filters.rest_framework import DjangoFilterBackend
from django_filters import FilterSet, RangeFilter, CharFilter, DateFromToRangeFilter
from . serializer import *
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from django.core.mail import EmailMessage, send_mail, get_connection
from django.conf import settings

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return result.getvalue()
    return None

class CustomPagination(PageNumberPagination):
    page_size = 5
    page_size_query_param = 'page_size'

class ItemRangeFilter(FilterSet):
    rate = RangeFilter()
    date = DateFromToRangeFilter()
    class Meta:
        model=Item
        fields=['rate', 'no', 'user', 'date']

class BillRangeFilter(FilterSet):
    amount = RangeFilter()
    date = DateFromToRangeFilter()
    items_id = CharFilter(lookup_expr='icontains', distinct= True)
    class Meta:
        model=Bill
        fields=['amount', 'user', 'items_id', 'date']

class CustomerRangeFilter(FilterSet):
    date = DateFromToRangeFilter()
    class Meta:
        model=Customer
        fields=['user', 'date']

# Create your views here.

class GoogleLogin(SocialLoginView):
    authentication_classes = []
    adapter_class = GoogleOAuth2Adapter
    callback_url = settings.FRONTEND_URL
    client_class = OAuth2Client

class FacebookLogin(SocialLoginView):
    authentication_classes = []
    adapter_class = FacebookOAuth2Adapter
    callback_url = settings.FRONTEND_URL
    client_class = OAuth2Client

class ImportDataView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    def create(self, request):
        data = request.data['dataToImport'].split('\n')
        id = 0
        page = request.data['page']
        try:
            with transaction.atomic():
                for d in data:
                    numbering = 1
                    if page=='products':
                        name = d.split('\t')[0]
                        rate = float(d.split('\t')[1])                    
                        while Item.objects.filter(user=request.data['user']).filter(name=name).exists():
                            name = d.split('\t')[0] + '_' + str(numbering)
                            numbering+=1
                        serializer = ItemSerializer(data = {'user': request.data['user'], 'no': str(int(request.data['no'])+id), 'name': name, 'rate': rate, 'date': request.data['date']})
                        if serializer.is_valid():
                            serializer.save()
                            id+=1
                        else:
                            raise Exception()
                    else:
                        name = d.split('\t')[0]
                        email = d.split('\t')[1]
                        phone = d.split('\t')[2]
                        if email:
                            if Customer.objects.filter(user=request.data['user']).filter(email=email).exists():
                                email = None
                        if phone:
                            if Customer.objects.filter(user=request.data['user']).filter(phone=phone).exists():
                                phone = None
                        serializer = CustomerSerializer(data = {'user': request.data['user'], 'no': str(int(request.data['no'])+id), 'name': name, 'email': email, 'phone': phone, 'date': request.data['date']})
                        if serializer.is_valid():
                            serializer.save()
                            id+=1
                        else:
                            raise Exception()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        except:
            return Response({'error': 'Data Error at row '+str(id+1)}, status=status.HTTP_404_NOT_FOUND)

class EmailConfirmationView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    def create (self, request):
        try:
            user = User.objects.get(id=request.data['user'])
            email = user.email
            user.is_active = False
            user.save()

            serializer = UserInfoSerializer(data = {'user':request.data['user'], 'no': request.data['no']})
            if serializer.is_valid():
                serializer.save()
            else:
                user.delete()
                return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)
            
            Token.objects.get(user = user).delete()
            token = Token.objects.get_or_create(user=user)[0]

            mail_subject = 'BillsApp - Email Verification'
            from_email = settings.EMAIL_HOST_USER
            to_email = [email]

            url = settings.FRONTEND_URL+"/account/verify/?user="+user.username+"&key="+str(token)
            message = get_template('verify_email.html').render({'username':user.username,'url':url})
            
            msg = EmailMessage(
                mail_subject,
                message,
                from_email,
                to_email,
            )
            msg.content_subtype = "html"
            msg.send()

            return Response(status=status.HTTP_200_OK)
        except :
            user.delete()
            return Response({'error': 'Error'}, status=status.HTTP_404_NOT_FOUND)

class CheckEmailConfirmationView(APIView):
    permission_classes = []
    def post(self, request):
        username = request.data['username']
        key = request.data['key']
        try:
            user = User.objects.get(username=username)
            token = Token.objects.get(user = user)
            if str(token)==str(key):
                user.is_active = True
                user.save()
                token.delete()
                return Response(status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)

class PasswordResetEmailView(APIView):
    permission_classes = []
    def post(self, request):
        email = request.data['email']
        try:
            user = User.objects.get(email=email)
        except :
            return Response({'error': 'Email ID not registered.'}, status=status.HTTP_404_NOT_FOUND)
        if user.is_active :
            username = user.username
            token = Token.objects.get_or_create(user = user)[0]
        else:
            return Response({'error': 'Email ID not verified yet.'}, status=status.HTTP_404_NOT_FOUND)
        
        mail_subject = 'BillsApp - Reset Password'
        from_email = settings.EMAIL_HOST_USER
        to_email = [email]

        url = settings.FRONTEND_URL+"/password-reset/?user="+username+"&key="+str(token)
        message = get_template('password_reset.html').render({'username':username,'url':url})
        
        msg = EmailMessage(
            mail_subject,
            message,
            from_email,
            to_email,
        )
        msg.content_subtype = "html"
        msg.send()

        return Response(status=status.HTTP_200_OK)

class PasswordResetCheckView(APIView):
    permission_classes = []
    def post(self, request):
        username = request.data['username']
        key = request.data['key']
        try:
            user = User.objects.get(username=username)
            token = Token.objects.get(user = user)
            if str(token)==str(key):
                token.delete()
                new_token = Token.objects.get_or_create(user = user)[0]
                return Response({'key': str(new_token), 'uid': user.id}, status=status.HTTP_200_OK)
            else:
                return Response(status=status.HTTP_404_NOT_FOUND)
        except:
            return Response(status=status.HTTP_404_NOT_FOUND)

class UpdatePasswordView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        try:
            validators.validate_password(password=request.data['password'])
            user = User.objects.get(id=request.data['uid'])
            user.set_password(request.data['password'])
            user.save()
            return Response(status=status.HTTP_200_OK)
        except exceptions.ValidationError as e:
            return Response({'error':e.messages}, status=status.HTTP_400_BAD_REQUEST)

class UserInfoCreateView(generics.CreateAPIView): 
    permission_classes = [IsAuthenticated]
    def create(self, request):
        user = User.objects.get(id=request.data['user'])
        old = UserInfo.objects.filter(user=user.id)
        if old.exists():
            serializer = UserInfoSerializer(old[0], data={'user': request.data['user'], 'avatar': request.data['avatar']})  
        else:
            serializer = UserInfoSerializer(data=request.data)
        
        if serializer.is_valid():
            serializer.save()
            return Response(status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)

class UserInfoGetView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = UserInfo.objects.all()
    serializer_class = UserInfoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['user']
 
class UserInfoUpdateView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = UserInfo.objects.all()
    serializer_class = UserInfoSerializer

class UserDelete(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        User.objects.filter(username=request.data['username']).delete()
        return Response(status=status.HTTP_200_OK)

class ItemCreateView(generics.CreateAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    def create(self, request):
        if request.data['status']=='save':
            serializer = ItemSerializer(data = {'user': request.data['user'], 'no': request.data['no'], 'name': request.data['name'], 'rate': request.data['rate'], 'date': request.data['date']})
            if serializer.is_valid():
                item = Item.objects.filter(user=request.data['user']).filter(name=request.data['name'])
                if item.exists() :
                    return Response({'name': 'Item with this name already exists.'}, status=status.HTTP_400_BAD_REQUEST)
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)
        elif request.data['status']=='duplicate':
            numbering = 1
            name = request.data['name']
            while Item.objects.filter(user=request.data['user']).filter(name=name).exists():
                name = request.data['name'] + '_' + str(numbering)
                numbering+=1
            serializer = ItemSerializer(data = {'user': request.data['user'], 'no': request.data['no'], 'name': name, 'rate': request.data['rate'], 'date': request.data['date']})
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            else:
                return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'name': 'Error'}, status=status.HTTP_404_NOT_FOUND)

class ItemGetView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    ordering_fields = ['no', 'name']
    filterset_fields = ['user', 'no', 'id']
    search_fields = ['name', 'no']

class ItemGetPageView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    pagination_class = CustomPagination
    filter_class = ItemRangeFilter
    ordering_fields = ['no', 'name', 'rate', 'date']
    search_fields = ['no', 'name', 'id']

class ItemUpdateView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    def patch(self, request, pk):
        obj = Item.objects.get(pk=pk)
        serializer = ItemSerializer(obj, data = request.data)
        if serializer.is_valid():
            item = Item.objects.filter(~Q(pk=pk)).filter(user=request.data['user']).filter(name=request.data['name'])
            if item.exists() :
                return Response({'name': 'Item with this name already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)

class ItemDeleteView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        Item.objects.all().delete()
        return Response(status=status.HTTP_200_OK)
    def post(self, request):
        Item.objects.filter(id__in = request.data['ids']).delete()
        return Response(status=status.HTTP_200_OK)

class BillCreateView(generics.CreateAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Bill.objects.all()
    serializer_class = BillSerializer

class BillGetView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['user', 'no', 'id']
    ordering_fields = ['no']

class BillGetPageView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Bill.objects.all()
    serializer_class = BillSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    pagination_class = CustomPagination
    filter_class = BillRangeFilter
    ordering_fields = ['id', 'no', 'customer_name', 'date', 'amount']
    search_fields = ['no', 'customer_name']

class BillUpdateView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Bill.objects.all()
    serializer_class = BillSerializer

class BillDeleteView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.query_params.get('userid')
        Bill.objects.all().delete()
        fs = FileSystemStorage()
        if os.path.isdir(os.path.join(settings.MEDIA_ROOT, 'saved\\'+user)):
            shutil.rmtree(os.path.join(settings.MEDIA_ROOT, 'saved\\'+user))
        return Response(status=status.HTTP_200_OK)
    def post(self, request):
        user = request.data['userid']
        bills = Bill.objects.filter(id__in = request.data['ids'])
        fs = FileSystemStorage()
        for bill in bills:
            for doc in bill.documents:
                fs.delete('saved/'+user+'/'+doc['name']);
        bills.delete()
        return Response(status=status.HTTP_200_OK)

class CustomerCreateView(generics.CreateAPIView): 
    permission_classes = [IsAuthenticated]
    def create(self, request):
        serializer = CustomerSerializer(data = request.data)
        if serializer.is_valid():
            if request.data['email']:
                customer1 = Customer.objects.filter(user=request.data['user']).filter(email=request.data['email'])
                if customer1.exists() :
                    return Response({'error': 'Customer with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST) 
            if request.data['phone']:
                customer2 = Customer.objects.filter(user=request.data['user']).filter(phone=request.data['phone'])
                if customer2.exists():
                    return Response({'error': 'Customer with this phone already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)

class CustomerGetView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ['user']
    search_fields = ['name', 'email']

class CustomerGetPageView(generics.ListAPIView): 
    permission_classes = [IsAuthenticated]
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter, filters.SearchFilter]
    pagination_class = CustomPagination
    filter_class = CustomerRangeFilter
    ordering_fields = ['id', 'name', 'date']
    search_fields = ['name', 'email', 'phone']

class CustomerUpdateView(generics.RetrieveUpdateDestroyAPIView):
    permission_classes = [IsAuthenticated]
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    def patch(self, request, pk):
        obj = Customer.objects.get(pk=pk)
        serializer = CustomerSerializer(obj, data = request.data)
        if serializer.is_valid():
            if request.data['email']:
                customer1 = Customer.objects.filter(~Q(pk=pk)).filter(user=request.data['user']).filter(email=request.data['email'])
                if customer1.exists():
                    return Response({'error': 'Customer with this email already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            if request.data['phone']:
                customer2 = Customer.objects.filter(~Q(pk=pk)).filter(user=request.data['user']).filter(phone=request.data['phone'])
                if customer2.exists():
                    return Response({'error': 'Customer with this phone already exists.'}, status=status.HTTP_400_BAD_REQUEST)
            
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        else:
            return Response(serializer.errors, status=status.HTTP_404_NOT_FOUND)

class CustomerDeleteView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        Customer.objects.all().delete()
        return Response(status=status.HTTP_200_OK)
    def post(self, request):
        Customer.objects.filter(id__in = request.data['ids']).delete()
        return Response(status=status.HTTP_200_OK)

class EmailFileSend(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        user = UserInfo.objects.get(id=request.data['user'])
        bill = Bill.objects.get(id=request.data['idToMail'])
        isRemarks = False
        if len(bill.remarks['data'])>7:
            isRemarks = True
        context = {
            'customer_name': bill.customer_name,
            'date': request.data['date'],
            'no': bill.id,
            'items': bill.items,
            'total_items': len(bill.items),
            'total_quantity': bill.quantity,
            'total_amount': bill.amount,
            'remarks': bill.remarks['data'],
            'isRemarks': isRemarks
        }

        pdf = render_to_pdf('invoice.html', context)

        if pdf:
            sender_address = user.email
            sender_pass = user.password
            receiver_address = request.data['to']
            message = MIMEMultipart()
            message['From'] = sender_address
            message['To'] = receiver_address
            message['Subject'] = request.data['subject']
            message.attach(MIMEText(request.data['body'], 'plain'))
            
            payload = MIMEBase('application', 'octate-stream', Name='bill_'+bill.no+'.pdf')
            payload.set_payload(pdf)
            encoders.encode_base64(payload)
            payload.add_header('Content-Disposition', 'attachment', filename='bill_'+bill.no+'.pdf')
            message.attach(payload)

            session = smtplib.SMTP(user.smtp_host, int(user.smtp_port))
            session.starttls()
            session.login(sender_address, sender_pass)
            text = message.as_string()
            session.sendmail(sender_address, receiver_address, text)
            session.quit()
        return Response(status=status.HTTP_200_OK)

class EmailSend(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        user = UserInfo.objects.get(id=request.data['user'])
        sender_address = user.email
        sender_pass = user.password
        receiver_address = request.data['to']
        message = MIMEMultipart()
        message['From'] = sender_address
        message['To'] = receiver_address
        message['Subject'] = request.data['subject']
        message.attach(MIMEText(request.data['body'], 'plain'))

        session = smtplib.SMTP(user.smtp_host, int(user.smtp_port))
        session.starttls()
        session.login(sender_address, sender_pass)
        text = message.as_string()
        session.sendmail(sender_address, receiver_address, text)
        session.quit()
        return Response(status=status.HTTP_200_OK)

class DocumentUploadView(APIView):
    permission_classes = [IsAuthenticated]
    parser_class = (MultiPartParser, FormParser)
    def post(self, request):
        fs = FileSystemStorage()
        filename = fs.save('unsaved/'+request.data['userid']+'/'+request.data['fileName'], request.data['file'])
        uploaded_file_url = fs.url(filename)
        return Response({'url': uploaded_file_url}, status=status.HTTP_200_OK)

class DocumentSaveView(APIView):
    permission_classes = [IsAuthenticated]
    def post(self, request):
        user = request.data['userid']
        names = request.data['names']
        fs = FileSystemStorage()
        if not os.path.isdir(os.path.join(settings.MEDIA_ROOT, 'saved\\'+user)):
            os.makedirs(os.path.join(settings.MEDIA_ROOT, 'saved\\'+user))
        for name in names:
            shutil.move(os.path.join(settings.MEDIA_ROOT, 'unsaved\\'+user+'\\'+name), os.path.join(settings.MEDIA_ROOT, 'saved\\'+user+'\\'+name))
        
        return Response(status = status.HTTP_200_OK)

class DocumentDeleteView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
        user = request.query_params.get('userid')
        fs = FileSystemStorage()
        if os.path.isdir(os.path.join(settings.MEDIA_ROOT, 'unsaved\\'+user)):
            shutil.rmtree(os.path.join(settings.MEDIA_ROOT, 'unsaved\\'+user))
        
        return Response(status = status.HTTP_200_OK)

    def post(self, request):
        user = request.data['userid']
        names = request.data['names']
        fs = FileSystemStorage()
        for name in names:
            fs.delete('unsaved/'+user+'/'+name);
            fs.delete('saved/'+user+'/'+name);
        return Response(status = status.HTTP_200_OK)

class DocumentDownloadView(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        url = request.query_params.get('url')
        folder, username, file_name = url.split('/')[-3:]
        file_path = settings.MEDIA_ROOT + '\\'+ folder +'\\'+ username +'\\' + file_name
        f = open(file_path, 'rb')
        file = File(f)
        response = HttpResponse(file.read())
        response['Content-Disposition'] = "attachment;filename=%s"%file_name
        return response

class ExportDataView(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request):
            if request.query_params.get('table')=='customers':
                query_set = Customer.objects.filter(user=request.query_params.get('user')).order_by(request.query_params.get('orderBy'))
                if request.query_params.get('searchQuery'):
                    query_set = query_set.filter(Q(name__icontains=request.query_params.get('searchQuery')) | Q(email__icontains=request.query_params.get('searchQuery')) | Q(phone__icontains=request.query_params.get('searchQuery')))
                if request.query_params.get('filterStartDate') or request.query_params.get('filterEndDate'):
                    query_set = query_set.filter(date__range=[request.query_params.get('filterStartDate'), request.query_params.get('filterEndDate')])
                data = json.loads(serialize('json', query_set))
                columns = ['Customer ID', 'Name', 'Email', 'Phone', 'Last Updated On']
                data_columns = ['no', 'name', 'email', 'phone', 'date']

            elif request.query_params.get('table')=='products':
                query_set = Item.objects.filter(user=request.query_params.get('user')).order_by(request.query_params.get('orderBy'))
                if request.query_params.get('searchQuery'):
                    query_set = query_set.filter(Q(no__icontains=request.query_params.get('searchQuery')) | Q(name__icontains=request.query_params.get('searchQuery')))
                if request.query_params.get('filterStartDate') or request.query_params.get('filterEndDate'):
                    query_set = query_set.filter(date__range=[request.query_params.get('filterStartDate'), request.query_params.get('filterEndDate')])
                if request.query_params.get('filterRateMin') or request.query_params.get('filterRateMax'):
                    query_set = query_set.filter(rate__gte=request.query_params.get('filterRateMin')).filter(rate__lte=request.query_params.get('filterRateMax'))
                data = json.loads(serialize('json', query_set))
                columns = ['Product ID', 'Name', 'Rate(Rs.)', 'Last Updated On']
                data_columns = ['no', 'name', 'rate', 'date']
            
            elif request.query_params.get('table')=='bills':
                query_set = Bill.objects.filter(user=request.query_params.get('user')).order_by(request.query_params.get('orderBy'))
                if request.query_params.get('searchQuery'):
                    query_set = query_set.filter(Q(customer_name__icontains=request.query_params.get('searchQuery')) | Q(no__icontains=request.query_params.get('searchQuery')))
                if request.query_params.get('filterProducts'):
                    query_set = query_set.filter(items_id__icontains=request.query_params.get('filterProducts'))
                if request.query_params.get('filterStartDate') or request.query_params.get('filterEndDate'):
                    query_set = query_set.filter(date__range=[request.query_params.get('filterStartDate'), request.query_params.get('filterEndDate')])
                if request.query_params.get('filterAmountMin') or request.query_params.get('filterAmountMax'):
                    query_set = query_set.filter(amount__gte=request.query_params.get('filterAmountMin')).filter(amount__lte=request.query_params.get('filterAmountMax'))
                data = json.loads(serialize('json', query_set))
                columns = ['Bill ID', 'Bill No.', 'Customer\'s Name', 'Customer\'s Email', 'Quantity', 'Amount(Rs.)', 'Last Updated On']
                data_columns = ['no', 'id', 'customer_name', 'customer_email', 'quantity', 'amount', 'date']

            print(data)
            rows = []
            for d in data:
                row=[]
                for col in data_columns:
                    if col=='date':
                        row.append(datetime.strptime(d['fields'][col], '%Y-%m-%dT%H:%M:%S.%f%z').strftime('%d %b, %Y - %I:%M:%S %p'))
                    elif col=='id':
                        row.append(d['pk'])
                    else:
                        row.append(d['fields'][col])
                rows.append(row)

            if request.query_params.get('type')=='xlsx':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename={table}_{date}.xlsx'.format(table=request.query_params.get('table'),date=str(int(datetime.now().timestamp())))

                workbook = Workbook()

                worksheet = workbook.active
                worksheet.title = request.query_params.get('table')+'_'+str(int(datetime.now().timestamp()))

                row_num = 1
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title
                    cell.font = Font(bold=True)
                    column_letter = get_column_letter(col_num)
                    column_dimensions = worksheet.column_dimensions[column_letter]
                    column_dimensions.width = 20

                for row in rows:
                    row_num+=1
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)

            else:
                pdf = render_to_pdf('table.html', {'heading': request.query_params.get('table').capitalize(), 'headers': columns, 'data': rows})
                response = HttpResponse(pdf)
                response['Content-Disposition'] = "attachment; filename={table}_{date}.pdf".format(table=request.query_params.get('table'),date=str(int(datetime.now().timestamp())))

            return response